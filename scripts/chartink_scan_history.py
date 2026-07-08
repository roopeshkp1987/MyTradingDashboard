"""
chartink_scan_history.py
========================
Runs a Chartink scan clause historically for the last 20 trading days
and saves the daily count of passing stocks to an Excel sheet.

Screener scan clause:
    ( {cash} ( n day ago close >= n day ago sma( daily close , 20 ) ) )

Requirements:
    pip install openpyxl requests beautifulsoup4 yfinance
"""

import argparse
import os
import sys
import time
from datetime import datetime
from pathlib import Path
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import yfinance as yf

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────
CHARTINK_URL = "https://chartink.com/screener/"
PROCESS_URL = "https://chartink.com/screener/process"

DEFAULT_HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/120.0.0.0 Safari/537.36'
    ),
    'Referer': 'https://chartink.com/screener/'
}

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def get_csrf_token(session: requests.Session) -> str:
    """Fetch the CSRF token from the Chartink main page."""
    resp = session.get(CHARTINK_URL, timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, 'html.parser')
    token_tag = soup.select_one("[name='csrf-token']")
    if not token_tag:
        raise ValueError("Could not find csrf-token in Chartink HTML response.")
    return token_tag.get('content', '')


def fetch_trading_dates(count: int = 20) -> list:
    """
    Fetch the last `count` trading dates for NSE (using NIFTY 50 as proxy).
    Returns list of datetime.date objects, newest first.
    """
    print(f"  Fetching last {count} trading dates from Yahoo Finance...")
    # Fetch 45 days to easily cover 20 trading days even with holidays/weekends
    ticker = "^NSEI"
    try:
        df = yf.download(ticker, period="45d", progress=False)
        if df.empty:
            raise ValueError("Empty response from Yahoo Finance.")
        # yfinance index contains Timestamp dates. Sort descending (newest first)
        dates = df.index.sort_values(ascending=False)
        trading_dates = [d.date() for d in dates[:count]]
        return trading_dates
    except Exception as e:
        print(f"  [Warning] Failed to fetch trading dates via yfinance: {e}")
        print("            Falling back to calendar weekdays...")
        # Fallback to last weekdays (does not skip Indian market holidays, but acts as best effort)
        import pandas as pd
        bdays = pd.bdate_range(end=datetime.today(), periods=count)[::-1]
        return [d.date() for d in bdays]


def run_scans_for_day(session: requests.Session, n: int) -> dict:
    """
    Execute 6 scan clauses on Chartink for `n` days ago.
    Returns a dictionary of counts.
    """
    scans = {
        "Above 20 SMA": f"( {{cash}} ( {n} day ago close >= {n} day ago sma( daily close , 20 ) ) )",
        "Below 20 SMA": f"( {{cash}} ( {n} day ago close < {n} day ago sma( daily close , 20 ) ) )",
        "Above 200 SMA": f"( {{cash}} ( {n} day ago close >= {n} day ago sma( daily close , 200 ) ) )",
        "Below 200 SMA": f"( {{cash}} ( {n} day ago close < {n} day ago sma( daily close , 200 ) ) )",
        "Move 3% Up": f"( {{cash}} ( {n} day ago \"close - 1 candle ago close / 1 candle ago close * 100\" >= 3.0 ) )",
        "Move 3% Down": f"( {{cash}} ( {n} day ago \"close - 1 candle ago close / 1 candle ago close * 100\" < -3.0 ) )"
    }
    
    counts = {}
    for name, scan_clause in scans.items():
        payload = {'scan_clause': scan_clause}
        resp = session.post(PROCESS_URL, data=payload, timeout=20)
        resp.raise_for_status()
        data = resp.json()
        counts[name] = data.get('recordsFiltered', len(data.get('data', [])))
        # Sleep to avoid hitting limits
        time.sleep(0.4)
        
    return counts


def write_to_excel(results: list, output_path: Path):
    """
    Write the date and count results to a styled Excel sheet.
    results: list of dicts [{'date': date_obj, 'day_ago': int, 'counts': dict}]
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Screener Results"
    
    # Grid lines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11, color="000000")
    stripe_fill = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
    
    thin_side = Side(border_style="thin", color="D3D3D3")
    border_style = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Header Row
    scan_names = [
        "Above 20 SMA", "Below 20 SMA", 
        "Above 200 SMA", "Below 200 SMA", 
        "Move 3% Up", "Move 3% Down"
    ]
    headers = ["Date", "Day Ago Index"] + scan_names
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style
    
    # Data Rows
    for row_idx, res in enumerate(results, 2):
        row_data = [
            res['date'].strftime('%Y-%m-%d'),
            res['day_ago']
        ]
        counts_dict = res.get('counts') or {}
        for name in scan_names:
            row_data.append(counts_dict.get(name, "Failed"))
            
        ws.append(row_data)
        
        # Apply styling to cells in the row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = data_font
            cell.border = border_style
            
            # Alignments
            if col_num == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col_num == 2:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
                
            # Striping for readability
            if row_idx % 2 == 0:
                cell.fill = stripe_fill
                
    # Column width auto-adjust
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
        
    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Run Chartink historical scan and save stock count to Excel."
    )
    parser.add_argument(
        "--days",
        type=int,
        default=20,
        help="Number of historical trading entries to generate (default: 20)"
    )
    parser.add_argument(
        "--out",
        type=str,
        default="chartink_scan_counts.xlsx",
        help="Output Excel file path (default: chartink_scan_counts.xlsx)"
    )
    
    args = parser.parse_args()
    
    output_path = Path(args.out)
    days_count = args.days
    
    print(f"\n{'='*65}")
    print(f"  Chartink Historical Scanner (6 Scans)")
    print(f"  Entries     : {days_count} days")
    print(f"  Output File : {output_path}")
    print(f"{'='*65}\n")
    
    # 1. Fetch Trading Dates
    trading_dates = fetch_trading_dates(days_count)
    if len(trading_dates) < days_count:
        print(f"  [Warning] Only fetched {len(trading_dates)} trading dates.")
        days_count = len(trading_dates)
        
    # 2. Init Session & fetch CSRF Token
    print("  Initializing Session & fetching CSRF Token...")
    session = requests.Session()
    session.headers.update(DEFAULT_HEADERS)
    
    try:
        csrf_token = get_csrf_token(session)
        session.headers.update({
            'x-csrf-token': csrf_token,
            'Content-Type': 'application/x-www-form-urlencoded'
        })
        print("  CSRF Token retrieved successfully.")
    except Exception as e:
        print(f"\n❌ Error initializing Chartink session: {e}")
        sys.exit(1)
        
    # 3. Query historical counts day by day
    print(f"\n  Querying 6 scans per day for {days_count} trading days...")
    results = []
    
    for n in range(days_count):
        date_obj = trading_dates[n]
        try:
            print(f"    [{n+1:2d}/{days_count}] Querying {date_obj.strftime('%Y-%m-%d')} ({n} day ago)... ", end="", flush=True)
            counts = run_scans_for_day(session, n)
            print(f"Done")
            results.append({
                'date': date_obj,
                'day_ago': n,
                'counts': counts
            })
            time.sleep(1.0)
        except Exception as e:
            print(f"FAILED: {e}")
            results.append({
                'date': date_obj,
                'day_ago': n,
                'counts': None
            })
            time.sleep(2)
            
    # 4. Save to Excel
    print(f"\n  Writing styled output to: {output_path.resolve()}")
    try:
        write_to_excel(results, output_path)
        print("  ✅ Excel sheet saved successfully!")
    except Exception as e:
        print(f"❌ Failed to save Excel file: {e}")
        sys.exit(1)
        
    # 5. Display Console Table Summary
    sep = "-" * 105
    scan_names = ["Above 20 SMA", "Below 20 SMA", "Above 200 SMA", "Below 200 SMA", "Move 3% Up", "Move 3% Down"]
    print(f"\n{sep}")
    header_str = f"  {'Date':<12} | {'Day Ago':<8} | " + " | ".join(f"{name:<13}" for name in scan_names)
    print(header_str)
    print(sep)
    for r in results:
        date_str = r['date'].strftime('%Y-%m-%d')
        day_ago = r['day_ago']
        c_dict = r.get('counts') or {}
        counts_str = " | ".join(f"{str(c_dict.get(name, 'Fail')):>13}" for name in scan_names)
        print(f"  {date_str:<12} | {day_ago:^8d} | {counts_str}")
    print(f"{sep}\n")


if __name__ == "__main__":
    main()
