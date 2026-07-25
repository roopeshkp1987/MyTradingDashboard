"""
NSE Market Breadth Scanner (Chartink-powered)
================================================
Runs 6 Chartink screeners and logs the resulting stock counts into an Excel
file - creating it if it doesn't exist, overwriting today's row if it's
already logged, or appending + re-sorting by date if it's a new day.

Screens run (via chartink.com's public screener engine):
   1. Close >= 21 EMA    (21 EMA Above)
   2. Close <  21 EMA    (21 EMA Below)
   3. Close >= 50 EMA    (50 EMA Above)
   4. Close <  50 EMA    (50 EMA Below)
   5. Close >= 200 EMA   (200 EMA Above)
   6. Close <  200 EMA   (200 EMA Below)
   7. Up   >= 3% for the day   (3% Up)
   8. Down <= -3% for the day  (3% Down)
   9. Up   >= 0% for the day   (Advance)
  10. Down <  0% for the day   (Decline)

Excel columns produced:
  A) SL#                              - serial number starting at 1
  B) Date                             - last NSE trading day (holiday-aware)
  C) 21 EMA Above
  D) 21 EMA Below
  E) 21 EMA Above/Below Ratio         (Excel formula, = C/D)
  F) 50 EMA Above
  G) 50 EMA Below
  H) 50 EMA Above/Below Ratio         (Excel formula, = F/G)
  I) 200 EMA Above
  J) 200 EMA Below
  K) 200 EMA Above/Below Ratio        (Excel formula, = I/J)
  L) 3% Up
  M) 3% Down
  N) 3% Up/Down Ratio                 (Excel formula, = L/M)
  O) Advance
  P) Decline
  Q) Advance/Decline Ratio            (Excel formula, = O/P)

IMPORTANT - about Chartink scan clauses
----------------------------------------
All clauses below use the "{cash} ( daily ... )" wrapping pattern, which
was confirmed working against the live chartink.com/screener/process
endpoint for the 3% up/down scans (and, by the same pattern, the Advance/
Decline scans you supplied). The EMA clauses (21/50/200 above/below) were
updated to match that same wrapping, but only the wrapping itself is
confirmed - the "daily close >= daily ema( daily close , N )" comparison
syntax inside it hasn't been independently verified live. If an EMA scan
returns 0 or an error, please run it once in the Chartink UI and paste the
exact scan_clause from the Network tab, the same way you did before.

Requirements:
    pip install requests openpyxl pandas_market_calendars --break-system-packages

    (pandas_market_calendars is optional but recommended - it lets the
    script correctly skip NSE holidays, not just weekends, when it stamps
    the "Date" column. Without it, the script falls back to a weekends-only
    calendar.)

Usage:
    python nse_breadth_scanner.py
    python nse_breadth_scanner.py --output my_log.xlsx
    python nse_breadth_scanner.py --delay 2.0     # slower, gentler on chartink
"""

import argparse
import re
import time
from datetime import datetime, timedelta
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    import pandas_market_calendars as mcal
    HAS_MCAL = True
except ImportError:
    HAS_MCAL = False


CHARTINK_SCREENER_PAGE = "https://chartink.com/screener"
CHARTINK_PROCESS_URL = "https://chartink.com/screener/process"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    ),
    "Origin": "https://chartink.com",
    "Referer": "https://chartink.com/screener",
}

# ---- Chartink scan clauses (all use the {cash} + daily wrapping pattern) ----
SCAN_CLAUSES = {
    "21ema_above":  '( {cash} ( daily close >= daily ema( close , 21 ) ) )',
    "21ema_below":  '( {cash} ( daily close < daily ema( close , 21 ) ) )',
    "50ema_above":  '( {cash} ( daily close >= daily ema( close , 50 ) ) )',
    "50ema_below":  '( {cash} ( daily close < daily ema( close , 50 ) ) )',
    "200ema_above": '( {cash} (  daily close >=  daily ema(  daily close , 200 ) ) )',
    "200ema_below": '( {cash} (  daily close <  daily ema(  daily close , 200) ) )',
    "up_3pct":      '( {cash} (  daily "close - 1 candle ago close / 1 candle ago close * 100" >=  3 ) )',
    "down_3pct":    '( {cash} (  daily "close - 1 candle ago close / 1 candle ago close * 100" <=  -3 ) )',
    "advance":      '( {cash} (  daily "close - 1 candle ago close / 1 candle ago close * 100" >=  0 ) )',
    "decline":      '( {cash} (  daily "close - 1 candle ago close / 1 candle ago close * 100" <  0 ) )',
}

SHEET_NAME = "Breadth Data"
COLUMNS = [
    "SL#", "Date",
    "21 EMA Above", "21 EMA Below", "21 EMA Above/Below Ratio",
    "50 EMA Above", "50 EMA Below", "50 EMA Above/Below Ratio",
    "200 EMA Above", "200 EMA Below", "200 EMA Above/Below Ratio",
    "3% Up", "3% Down", "3% Up/Down Ratio",
    "Advance", "Decline", "Advance/Decline Ratio",
]
COLUMN_WIDTHS = [6, 13, 14, 14, 16, 14, 14, 16, 14, 14, 17, 10, 10, 14, 10, 10, 17]


# ---------------- Chartink session / scanning ----------------
def get_csrf_session():
    """Opens a session against chartink.com and pulls the CSRF token needed
    for POSTing to the screener/process endpoint."""
    session = requests.Session()
    session.headers.update(HEADERS)
    resp = session.get(CHARTINK_SCREENER_PAGE, timeout=15)
    resp.raise_for_status()
    match = re.search(r'name="csrf-token"\s+content="([^"]+)"', resp.text)
    if not match:
        raise RuntimeError(
            "Could not find a csrf-token on the chartink screener page. "
            "The site's markup may have changed - inspect the page source "
            "for the current meta tag name."
        )
    session.headers.update({"x-csrf-token": match.group(1)})
    return session


def run_scan(session, clause, retries=3, backoff=2.0):
    """POSTs one scan clause to chartink and returns the matching stock rows."""
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            resp = session.post(
                CHARTINK_PROCESS_URL, data={"scan_clause": clause}, timeout=20
            )
            resp.raise_for_status()
            payload = resp.json()
            return payload.get("data", [])
        except (requests.RequestException, ValueError) as exc:
            last_err = exc
            if attempt < retries:
                time.sleep(backoff * attempt)
    raise RuntimeError(f"Chartink scan failed after {retries} attempts: {last_err}")


def compute_breadth(session, delay=1.5):
    """Runs all 10 screens and returns a dict of stock counts."""
    counts = {}
    for key, clause in SCAN_CLAUSES.items():
        print(f"  Running scan: {key} ...")
        rows = run_scan(session, clause)
        counts[key] = len(rows)
        time.sleep(delay)  # be polite to chartink's servers
    return counts


# ---------------- NSE trading-day resolution (holiday-aware) ----------------
def get_last_trading_day(as_of=None):
    """Returns the most recent NSE trading day on/before `as_of`.

    Uses pandas_market_calendars' NSE-India calendar ("XNSE") when available
    so real NSE holidays are skipped, not just weekends. Falls back to a
    weekends-only check if that package isn't installed.
    """
    as_of = as_of or datetime.now().date()

    if HAS_MCAL:
        cal = mcal.get_calendar("XNSE")
        schedule = cal.schedule(start_date=as_of - timedelta(days=14), end_date=as_of)
        if not schedule.empty:
            return schedule.index[-1].date()

    # Fallback: weekends-only, no holiday awareness
    d = as_of
    while d.weekday() >= 5:  # Saturday=5, Sunday=6
        d -= timedelta(days=1)
    return d


# ---------------- Excel output (create / upsert-by-date / re-sort) ----------------
def _style_header(ws):
    header_fill = PatternFill(start_color="1F2933", end_color="1F2933", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, w in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _to_date(val):
    """Normalizes a date/datetime/string cell value to a plain date object."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if hasattr(val, "isoformat") and not isinstance(val, str):
        return val
    if isinstance(val, str):
        return datetime.strptime(val[:10], "%Y-%m-%d").date()
    return val


def read_existing_rows(ws):
    """Reads all existing data rows (below the header) into plain dicts."""
    rows = []
    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(row=r, column=2).value
        d = _to_date(date_val)
        if d is None:
            continue
        rows.append({
            "Date": d,
            "21EMA_Above": ws.cell(row=r, column=3).value,
            "21EMA_Below": ws.cell(row=r, column=4).value,
            "50EMA_Above": ws.cell(row=r, column=6).value,
            "50EMA_Below": ws.cell(row=r, column=7).value,
            "200EMA_Above": ws.cell(row=r, column=9).value,
            "200EMA_Below": ws.cell(row=r, column=10).value,
            "Up_3pct": ws.cell(row=r, column=12).value,
            "Down_3pct": ws.cell(row=r, column=13).value,
            "Advance": ws.cell(row=r, column=15).value,
            "Decline": ws.cell(row=r, column=16).value,
        })
    return rows


def upsert(rows, new_entry):
    """Overwrites the row for new_entry['Date'] if present, else appends."""
    for i, r in enumerate(rows):
        if r["Date"] == new_entry["Date"]:
            rows[i] = new_entry
            return rows, "Overwrote"
    rows.append(new_entry)
    return rows, "Appended"


def write_result(new_entry, output_path):
    output_path = Path(output_path)

    if output_path.exists():
        wb = openpyxl.load_workbook(output_path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        rows = read_existing_rows(ws)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        rows = []

    rows, action = upsert(rows, new_entry)
    rows.sort(key=lambda r: r["Date"])  # REQ-3: keep the sheet sorted by date

    # Wipe and rebuild the sheet body so SL# and formula row-refs stay correct
    ws.delete_rows(1, ws.max_row)
    _style_header(ws)

    input_font = Font(name="Arial", color="0000FF", size=11)
    calc_font = Font(name="Arial", italic=True, color="006100", size=11)

    for idx, r in enumerate(rows, start=1):
        row_num = idx + 1

        ws.cell(row=row_num, column=1, value=idx).font = input_font

        date_cell = ws.cell(row=row_num, column=2, value=r["Date"])
        date_cell.number_format = "yyyy-mm-dd"
        date_cell.font = input_font

        ws.cell(row=row_num, column=3, value=r["21EMA_Above"]).font = input_font
        ws.cell(row=row_num, column=4, value=r["21EMA_Below"]).font = input_font
        e_cell = ws.cell(row=row_num, column=5, value=f"=IFERROR(C{row_num}/D{row_num},0)")
        e_cell.font = calc_font
        e_cell.number_format = "0.00"

        ws.cell(row=row_num, column=6, value=r["50EMA_Above"]).font = input_font
        ws.cell(row=row_num, column=7, value=r["50EMA_Below"]).font = input_font
        h_cell = ws.cell(row=row_num, column=8, value=f"=IFERROR(F{row_num}/G{row_num},0)")
        h_cell.font = calc_font
        h_cell.number_format = "0.00"

        ws.cell(row=row_num, column=9, value=r["200EMA_Above"]).font = input_font
        ws.cell(row=row_num, column=10, value=r["200EMA_Below"]).font = input_font
        k_cell = ws.cell(row=row_num, column=11, value=f"=IFERROR(I{row_num}/J{row_num},0)")
        k_cell.font = calc_font
        k_cell.number_format = "0.00"

        ws.cell(row=row_num, column=12, value=r["Up_3pct"]).font = input_font
        ws.cell(row=row_num, column=13, value=r["Down_3pct"]).font = input_font
        n_cell = ws.cell(row=row_num, column=14, value=f"=IFERROR(L{row_num}/M{row_num},0)")
        n_cell.font = calc_font
        n_cell.number_format = "0.00"

        ws.cell(row=row_num, column=15, value=r["Advance"]).font = input_font
        ws.cell(row=row_num, column=16, value=r["Decline"]).font = input_font
        q_cell = ws.cell(row=row_num, column=17, value=f"=IFERROR(O{row_num}/P{row_num},0)")
        q_cell.font = calc_font
        q_cell.number_format = "0.00"

        for col in range(1, len(COLUMNS) + 1):
            ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="center")

    wb.save(output_path)
    print(f"{action} row for {new_entry['Date']} - sheet now has {len(rows)} row(s), sorted by date.")


# ---------------- Main ----------------

def main():
    parser = argparse.ArgumentParser(description="NSE market breadth scanner (Chartink-powered)")
    parser.add_argument("--output", type=str, default="nse_breadth_data.xlsx", help="Output Excel filename")
    parser.add_argument("--delay", type=float, default=1.5, help="Seconds to wait between chartink scan calls")
    args = parser.parse_args()

    print("Connecting to Chartink and fetching a CSRF token...")
    session = get_csrf_session()

    print("Running screeners...")
    counts = compute_breadth(session, delay=args.delay)

    trading_day = get_last_trading_day()

    new_entry = {
        "Date": trading_day,
        "21EMA_Above": counts["21ema_above"],
        "21EMA_Below": counts["21ema_below"],
        "50EMA_Above": counts["50ema_above"],
        "50EMA_Below": counts["50ema_below"],
        "200EMA_Above": counts["200ema_above"],
        "200EMA_Below": counts["200ema_below"],
        "Up_3pct": counts["up_3pct"],
        "Down_3pct": counts["down_3pct"],
        "Advance": counts["advance"],
        "Decline": counts["decline"],
    }

    output_path = Path.cwd() / args.output
    write_result(new_entry, output_path)

    print("\nToday's breadth (Chartink):")
    print(f"  Date:          {new_entry['Date']}")
    print(f"  21EMA Above:   {new_entry['21EMA_Above']}")
    print(f"  21EMA Below:   {new_entry['21EMA_Below']}")
    print(f"  50EMA Above:   {new_entry['50EMA_Above']}")
    print(f"  50EMA Below:   {new_entry['50EMA_Below']}")
    print(f"  200EMA Above:  {new_entry['200EMA_Above']}")
    print(f"  200EMA Below:  {new_entry['200EMA_Below']}")
    print(f"  3% Up:         {new_entry['Up_3pct']}")
    print(f"  3% Down:       {new_entry['Down_3pct']}")
    print(f"  Advance:       {new_entry['Advance']}")
    print(f"  Decline:       {new_entry['Decline']}")


if __name__ == "__main__":
    main()